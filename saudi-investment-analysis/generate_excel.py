"""
Saudi Investment Portfolio Analysis - Excel Generator
Generates comprehensive Excel workbook with 9 Saudi companies analysis
Data sourced from: Saudi Exchange (Tadawul), Argaam, Mubasher (May 2026)

Run `python3 fetch_prices.py` to auto-update prices before generating.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import json
import os


def _load_prices_from_file() -> dict:
    """يحمّل الأسعار من prices.json إذا كان موجوداً ومحدَّثاً."""
    path = os.path.join(os.path.dirname(__file__), "prices.json")
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        prices = {k: v for k, v in data.items() if not k.startswith("_") and isinstance(v, (int, float))}
        updated = data.get("_last_updated", "")
        if prices:
            print(f"📌 أسعار محمَّلة من prices.json (آخر تحديث: {updated})")
        return prices
    except Exception:
        return {}

_LIVE_PRICES = _load_prices_from_file()

# ─── COLOR PALETTE ────────────────────────────────────────────────────────────
C_DARK_NAVY   = "1B2A4A"
C_GOLD        = "C9A84C"
C_LIGHT_BLUE  = "D6E4F0"
C_WHITE       = "FFFFFF"
C_HEADER_BG   = "1B2A4A"
C_SUB_HEADER  = "2E4172"
C_POSITIVE    = "1A7F37"
C_NEGATIVE    = "C0392B"
C_WARNING     = "E67E22"
C_NEUTRAL     = "2980B9"
C_LIGHT_GREEN = "D5F5E3"
C_LIGHT_RED   = "FADBD8"
C_LIGHT_GRAY  = "F2F3F4"
C_BLEND       = "EAF2FF"
C_GROWTH      = "FFF3E0"
C_VALUE       = "F0FFF0"

def make_border(color="CCCCCC"):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=11, bold=True, color=C_WHITE):
    return Font(name='Calibri', size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name='Calibri', size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def right_align():
    return Alignment(horizontal='right', vertical='center')

def left_align():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

# ─── COMPANY DATA ─────────────────────────────────────────────────────────────
COMPANIES = {
    "الراجحي": {
        "ticker": "1120",
        "sector": "مصرفي",
        "classification": "Blend",
        "shares_m": 6000,
        "capital_m": 60000,
        "price_sar": 69.35,
        "note_capital": "✅ بعد منح الأسهم (1 سهم مجاني لكل 2) → الأسهم أصبحت 6B سهم | سعر مُعدَّل",
        "profits_m": {2021: 14979, 2022: 17151, 2023: 16621, 2024: 19722},
        "revenues_m": {2021: 23500, 2022: 30100, 2023: 29800, 2024: 34200},
        "op_cashflow_m": {2021: 12000, 2022: 16000, 2023: 14500, 2024: 18500},
        "equity_m": {2021: 62000, 2022: 72000, 2023: 90259, 2024: 100185},
        "divs": [
            ("H1 2020", "12.5%", 1.25, 5000, None, None),
            ("H2 2020", "10.0%", 1.00, 4000, None, None),
            ("H1 2021", "12.5%", 1.25, 5000, None, None),
            ("H2 2021", "9.375%", 0.9375, 3750, None, None),
            ("H1 2022", "12.5%", 1.25, 5000, None, None),
            ("H2 2022", "12.5%", 1.25, 5000, None, None),
            ("H1 2023", "11.5%", 1.15, 4600, None, None),
            ("H2 2023", "11.5%", 1.15, 4600, None, None),
            ("H1 2024", "12.5%", 1.25, 5000, None, None),
            ("H2 2024", "14.6%", 1.46, 5840, None, None),
            ("H1 2025", "7.5%",  0.75, 3000, None, None),
            ("H2 2025", "17.5%", 1.75, 7000, "⚠ قبل المنح", None),
        ],
        "analysts": {"min": 75, "avg": 88, "max": 102},
        "pe": 14.0,
        "pb": 2.76,
        "roe_pct": 20.2,
        "roa_pct": 2.1,
        "div_yield_pct": 3.9,
        "fair_value": {"pe_based": 83, "ddm": 78, "pb_based": 90},
        "vision2030": "تمويل مشاريع عملاقة + صيرفة إسلامية + تمويل الأفراد + رقمنة الخدمات",
        "mgmt_score": 4,
        "risk_level": "منخفض-متوسط",
        "key_risks": "متطلبات كفاية رأس المال (SAMA) · تأثير المنح على المعدلات للسهم · تنافسية ضيقة",
        "recommendation": "تراكم",
    },
    "الإنماء": {
        "ticker": "1150",
        "sector": "مصرفي",
        "classification": "Growth",
        "shares_m": 1992,
        "capital_m": 19922,
        "price_sar": 23.86,
        "note_capital": "تحول إلى توزيعات ربعية منذ 2023",
        "profits_m": {2021: 2714, 2022: 3599, 2023: 4839, 2024: 5832},
        "revenues_m": {2021: 5500, 2022: 7200, 2023: 8700, 2024: 10200},
        "op_cashflow_m": {2021: 3000, 2022: 4200, 2023: 5500, 2024: 6800},
        "equity_m": {2021: 18000, 2022: 21000, 2023: 25000, 2024: 29500},
        "divs": [
            ("Q1 2023", "2.5%", 0.25, 498, None, None),
            ("Q2 2023", "2.5%", 0.25, 498, None, None),
            ("Q3 2023", "2.5%", 0.25, 498, None, None),
            ("Q4 2023", "2.5%", 0.25, 498, None, None),
            ("Q1 2024", "2.5%", 0.25, 498, None, None),
            ("Q2 2024", "2.5%", 0.25, 498, None, None),
            ("Q3 2024", "3.0%", 0.30, 598, None, None),
            ("Q4 2024", "3.0%", 0.30, 598, None, None),
            ("Q1 2025", "3.0%", 0.30, 598, None, None),
            ("Q2 2025", "3.0%", 0.30, 598, None, None),
            ("Q3 2025", "3.0%", 0.30, 598, None, None),
            ("Q1 2026", "2.5%", 0.25, 498, None, None),
        ],
        "analysts": {"min": 25, "avg": 31, "max": 38},
        "pe": 9.4,
        "pb": 1.86,
        "roe_pct": 21.3,
        "roa_pct": 2.0,
        "div_yield_pct": 4.4,
        "fair_value": {"pe_based": 35, "ddm": 28, "pb_based": 32},
        "vision2030": "تمويل المشاريع الكبرى + تمويل الرهن العقاري + SME Lending + رقمنة",
        "mgmt_score": 4,
        "risk_level": "منخفض-متوسط",
        "key_risks": "بنك أصغر نسبياً · انكشاف عقاري · سيولة تداول أقل",
        "recommendation": "تراكم",
    },
    "STC": {
        "ticker": "7010",
        "sector": "اتصالات",
        "classification": "Blend",
        "shares_m": 5000,
        "capital_m": 50000,
        "price_sar": 43.54,
        "note_capital": "⚠ ربح 2024 (24.7B) يشمل بيع TAWAL (13.97B غير متكرر) → الربح الجاري ≈ 10.7B ريال",
        "profits_m": {2021: 11100, 2022: 12200, 2023: 13300, 2024: 24689},
        "revenues_m": {2021: 63000, 2022: 67000, 2023: 71800, 2024: 75893},
        "op_cashflow_m": {2021: 18000, 2022: 20000, 2023: 22000, 2024: 24500},
        "equity_m": {2021: 50000, 2022: 52000, 2023: 48000, 2024: 55000},
        "divs": [
            ("Q1 2022", "4.0%", 0.40, 2000, None, None),
            ("Q2 2022", "4.0%", 0.40, 2000, None, None),
            ("Q3 2022", "4.0%", 0.40, 2000, None, None),
            ("Q4 2022", "4.0%", 0.40, 2000, None, None),
            ("Q1 2023", "4.0%", 0.40, 2000, None, None),
            ("Q2 2023", "4.0%", 0.40, 2000, None, None),
            ("Q3 2023", "4.0%", 0.40, 2000, None, None),
            ("Q4 2023", "4.0%", 0.40, 2000, None, None),
            ("Q1 2024", "4.0%", 0.40, 2000, None, None),
            ("Q2 2024", "4.0%", 0.40, 2000, None, None),
            ("Q3 2024", "4.0%", 0.40, 2000, None, None),
            ("Q4 2024 + استثنائي", "25.5%", 2.55, 12750, "⚠ يشمل 2 ريال بند غير متكرر", None),
            ("Q1 2025", "5.5%", 0.55, 2750, None, None),
            ("Q2 2025", "5.5%", 0.55, 2750, None, None),
            ("Q3 2025", "5.5%", 0.55, 2750, None, None),
            ("Q4 2025", "5.5%", 0.55, 2750, None, None),
        ],
        "analysts": {"min": 38, "avg": 47, "max": 56},
        "pe": 14.2,
        "pb": 3.82,
        "roe_pct": 26.9,
        "roa_pct": 4.2,
        "div_yield_pct": 5.2,
        "fair_value": {"pe_based": 51, "ddm": 45, "pb_based": 48},
        "vision2030": "البنية الرقمية الوطنية + 5G + STC Pay + حلول سحابة للقطاع الحكومي",
        "mgmt_score": 4,
        "risk_level": "منخفض",
        "key_risks": "CAPEX مرتفع · ملكية حكومية قد تؤثر على قرارات التوزيع · منافسة موبايلي/زين",
        "recommendation": "تراكم بقوة",
    },
    "سال": {
        "ticker": "4263",
        "sector": "لوجستيات",
        "classification": "Growth",
        "shares_m": 80,
        "capital_m": 800,
        "price_sar": 168.00,
        "note_capital": "IPO حديث (2022) - بيانات تاريخية محدودة | ✅ إيرادات 2024: 1.63B ريال (تحقق)",
        "profits_m": {2021: 320, 2022: 450, 2023: 511, 2024: 661},
        "revenues_m": {2021: 900, 2022: 1100, 2023: 1452, 2024: 1630},
        "op_cashflow_m": {2021: 400, 2022: 550, 2023: 680, 2024: 820},
        "equity_m": {2021: 1200, 2022: 1500, 2023: 1900, 2024: 2300},
        "divs": [
            ("Q4 2022", "5.0%", 0.50, 40, None, None),
            ("Q4 2023", "10.0%", 1.00, 80, None, None),
            ("Q4 2024", "13.3%", 1.33, 106, "✅ تحقق من تداول", None),
            ("Q1 2025", "14.3%", 1.43, 114, "✅ تحقق: 1.43 وليس 0.75", None),
            ("Q2 2025", "15.2%", 1.52, 122, "✅ تحقق", None),
            ("Q3 2025", "17.0%", 1.70, 136, "✅ تحقق", None),
        ],
        "analysts": {"min": 160, "avg": 205, "max": 250},
        "pe": 22.4,
        "pb": 6.4,
        "roe_pct": 30.5,
        "roa_pct": 8.2,
        "div_yield_pct": 1.8,
        "fair_value": {"pe_based": 220, "ddm": 195, "pb_based": 210},
        "vision2030": "مطار الملك سلمان الدولي + نمو الشحن الجوي + خدمات الخطوط السعودية",
        "mgmt_score": 4,
        "risk_level": "متوسط",
        "key_risks": "تركيز عملاء عالٍ · IPO تقييم مرتفع · بيانات تاريخية قصيرة · تذبذب الأسعار",
        "recommendation": "مراقبة / تراكم تدريجي",
    },
    "المواساة": {
        "ticker": "4002",
        "sector": "رعاية صحية",
        "classification": "Value",
        "shares_m": 200,
        "capital_m": 2000,
        "price_sar": 59.10,
        "note_capital": "توزيع سنوي ثابت منذ الإدراج",
        "profits_m": {2021: 520, 2022: 680, 2023: 658, 2024: 646},
        "revenues_m": {2021: 2050, 2022: 2400, 2023: 2706, 2024: 2879},
        "op_cashflow_m": {2021: 650, 2022: 820, 2023: 800, 2024: 780},
        "equity_m": {2021: 2800, 2022: 3200, 2023: 3600, 2024: 4000},
        "divs": [
            ("2020", "10.0%", 1.00, 200, None, None),
            ("2021", "12.5%", 1.25, 250, None, None),
            ("2022", "30.0%", 3.00, 600, "توزيع استثنائي مرتفع", None),
            ("2023", "17.5%", 1.75, 350, None, None),
            ("2024", "20.0%", 2.00, 400, None, None),
            ("H2 2025", "11.25%", 1.125, 225, "نصف سنوي", None),
        ],
        "analysts": {"min": 62, "avg": 78, "max": 92},
        "pe": 20.7,
        "pb": 3.35,
        "roe_pct": 17.3,
        "roa_pct": 6.8,
        "div_yield_pct": 3.0,
        "fair_value": {"pe_based": 74, "ddm": 70, "pb_based": 80},
        "vision2030": "خصخصة الصحة + التأمين الإلزامي + السياحة الطبية + توسع في المدينة",
        "mgmt_score": 3,
        "risk_level": "منخفض",
        "key_risks": "تراجع هامش الربح · شح الكوادر الطبية · تنافسية القطاع الخاص المتنامية",
        "recommendation": "احتفاظ / تراكم عند الضعف",
    },
    "بدجت": {
        "ticker": "4260",
        "sector": "خدمات استهلاكية",
        "classification": "Growth",
        "shares_m": 100,
        "capital_m": 1000,
        "price_sar": 75.90,
        "note_capital": "⚠ انخفاض أرباح Q1 2026 بنسبة 58% - يستوجب مراقبة",
        "profits_m": {2021: 185, 2022: 240, 2023: 277, 2024: 312},
        "revenues_m": {2021: 1200, 2022: 1500, 2023: 1750, 2024: 1980},
        "op_cashflow_m": {2021: 380, 2022: 460, 2023: 510, 2024: 550},
        "equity_m": {2021: 1300, 2022: 1500, 2023: 1725, 2024: 1903},
        "divs": [
            ("H2 2021", "5.0%", 0.50, 50, None, None),
            ("H1 2022", "5.0%", 0.50, 50, None, None),
            ("H2 2022", "5.0%", 0.50, 50, None, None),
            ("H1 2023", "5.0%", 0.50, 50, "✅ تحقق: أغسطس 2023", None),
            ("H2 2023", "10.0%", 1.00, 100, "✅ تحقق: مايو 2024 من أرباح FY2023", None),
            ("H1 2024", "4.5%", 0.45, 45, "✅ تحقق: أغسطس 2024", None),
            ("H2 2024", "10.0%", 1.00, 100, "✅ تحقق: مايو 2025 من أرباح FY2024", None),
            ("H1 2025", "5.0%", 0.50, 50, None, None),
        ],
        "analysts": {"min": 32, "avg": 42, "max": 52},
        "pe": 12.2,
        "pb": 2.0,
        "roe_pct": 17.4,
        "roa_pct": 5.1,
        "div_yield_pct": 3.8,
        "fair_value": {"pe_based": 46, "ddm": 38, "pb_based": 40},
        "vision2030": "السياحة + الفعاليات والترفيه + تأجير أسطول NEOM + نمو المطارات",
        "mgmt_score": 3,
        "risk_level": "متوسط-مرتفع",
        "key_risks": "⚠ تراجع ربح Q1 2026 بـ58% · رأس مال مكثف للأسطول · منافسة منصات المشاركة",
        "recommendation": "مراقبة - انتظار توضيح Q1 2026",
    },
    "اكسترا": {
        "ticker": "4003",
        "sector": "تجزئة",
        "classification": "Value",
        "shares_m": 80,
        "capital_m": 800,
        "price_sar": 81.90,
        "note_capital": "✅ توزيع استثنائي أغسطس 2024: 5 ريال/سهم من الاحتياطيات | ربح FY2024: 534.5M ريال",
        "profits_m": {2021: 280, 2022: 340, 2023: 390, 2024: 534},
        "revenues_m": {2021: 4500, 2022: 5200, 2023: 5800, 2024: 6781},
        "op_cashflow_m": {2021: 350, 2022: 430, 2023: 500, 2024: 620},
        "equity_m": {2021: 1200, 2022: 1400, 2023: 1650, 2024: 1900},
        "divs": [
            ("H1 2021", "10.0%", 1.00, 80, None, None),
            ("H2 2021", "10.0%", 1.00, 80, None, None),
            ("H1 2022", "15.0%", 1.50, 120, None, None),
            ("H2 2022", "15.0%", 1.50, 120, None, None),
            ("H1 2023", "15.0%", 1.50, 120, None, None),
            ("H2 2023", "15.0%", 1.50, 120, None, None),
            ("H1 2024", "20.0%", 2.00, 160, None, None),
            ("H1 2024 استثنائي", "50.0%", 5.00, 400, "✅ تحقق: توزيع استثنائي أغسطس 2024 من الاحتياطيات", None),
            ("H2 2024", "30.0%", 3.00, 240, "✅ تحقق", None),
            ("H1 2025", "20.0%", 2.00, 160, None, None),
        ],
        "analysts": {"min": 85, "avg": 107, "max": 130},
        "pe": 14.4,
        "pb": 4.0,
        "roe_pct": 30.1,
        "roa_pct": 11.2,
        "div_yield_pct": 5.3,
        "fair_value": {"pe_based": 115, "ddm": 98, "pb_based": 105},
        "vision2030": "نمو الإنفاق الاستهلاكي + التحول الرقمي + تراخيص جديدة + توسع جغرافي",
        "mgmt_score": 4,
        "risk_level": "منخفض-متوسط",
        "key_risks": "منافسة أمازون.sa ونون · ضغط هوامش التجزئة · التحدي الرقمي",
        "recommendation": "تراكم",
    },
    "المتقدمة": {
        "ticker": "2330",
        "sector": "بتروكيماويات",
        "classification": "Blend",
        "shares_m": 125,
        "capital_m": 1250,
        "price_sar": 34.12,
        "note_capital": "⚠ لا توزيعات 2023-2024 بسبب الخسارة | العودة للربحية Q1 2025",
        "profits_m": {2021: 380, 2022: 320, 2023: 85, 2024: -259},
        "revenues_m": {2021: 2100, 2022: 2400, 2023: 1900, 2024: 1700},
        "op_cashflow_m": {2021: 450, 2022: 380, 2023: 180, 2024: 50},
        "equity_m": {2021: 2200, 2022: 2400, 2023: 2300, 2024: 2000},
        "divs": [
            ("Q2 2022", "4.0%", 0.40, 50, None, None),
            ("Q4 2022", "4.0%", 0.40, 50, None, None),
            ("2023", "لا توزيع", 0, 0, "خسارة", None),
            ("2024", "لا توزيع", 0, 0, "خسارة", None),
            ("2025", "مفوض*", None, None, "الجمعية فوضت مجلس الإدارة", None),
        ],
        "analysts": {"min": 42, "avg": 56, "max": 70},
        "pe": None,
        "pb": 3.0,
        "roe_pct": None,
        "roa_pct": None,
        "div_yield_pct": 0.0,
        "fair_value": {"pe_based": 60, "ddm": None, "pb_based": 55},
        "vision2030": "تحويل البلاستيك للمواد الإنشائية (NEOM) + طاقة انتاجية جديدة + تنويع المنتجات",
        "mgmt_score": 3,
        "risk_level": "مرتفع",
        "key_risks": "⚠ دورية البتروكيماويات · خسارة 2024 · أسعار البروبيلين العالمية · منافسة صينية",
        "recommendation": "مضاربة / وزن محدود",
    },
    "بنيان ريت": {
        "ticker": "4340",
        "sector": "صندوق عقاري (REIT)",
        "classification": "Value",
        "shares_m": 163,
        "capital_m": 1630,
        "price_sar": 9.25,
        "note_capital": "توزيع نصف سنوي إلزامي ≥90% من الدخل الصافي",
        "profits_m": {2021: 85, 2022: 95, 2023: 105, 2024: 118},
        "revenues_m": {2021: 140, 2022: 155, 2023: 170, 2024: 190},
        "op_cashflow_m": {2021: 90, 2022: 100, 2023: 115, 2024: 130},
        "equity_m": {2021: 1550, 2022: 1570, 2023: 1590, 2024: 1610},
        "divs": [
            ("مايو-أكتوبر 2022", "3.0%", 0.30, 48.9, None, None),
            ("نوف 2022 - أبريل 2023", "3.0%", 0.30, 48.9, None, None),
            ("مايو-أكتوبر 2023", "3.2%", 0.32, 52.2, None, None),
            ("نوف 2023 - أبريل 2024", "3.2%", 0.32, 52.2, None, None),
            ("مايو-أكتوبر 2024", "3.4%", 0.34, 55.4, None, None),
            ("نوف 2024 - أبريل 2025", "3.7%", 0.37, 60.3, None, None),
            ("مايو-أكتوبر 2025", "4.0%", 0.40, 65.2, None, None),
        ],
        "analysts": {"min": 9.5, "avg": 11.2, "max": 13.0},
        "pe": 14.1,
        "pb": 1.03,
        "roe_pct": 7.3,
        "roa_pct": 5.2,
        "div_yield_pct": 7.5,
        "fair_value": {"pe_based": 10.8, "ddm": 11.5, "pb_based": 10.5},
        "vision2030": "نمو القطاع العقاري + الطلب السياحي + نمو الإيجارات التجارية",
        "mgmt_score": 3,
        "risk_level": "منخفض-متوسط",
        "key_risks": "حساسية أسعار الفائدة · جودة المستأجرين · إعادة التقييم العقاري · تركيز الأصول",
        "recommendation": "احتفاظ للدخل",
    },
}

COMPANY_ORDER = ["الراجحي", "الإنماء", "STC", "سال", "المواساة", "بدجت", "اكسترا", "المتقدمة", "بنيان ريت"]

# تطبيق الأسعار من prices.json (إن وُجدت) على بيانات الشركات
if _LIVE_PRICES:
    for _name, _price in _LIVE_PRICES.items():
        if _name in COMPANIES and _price > 0:
            COMPANIES[_name]["price_sar"] = _price

# ─── ANALYST CONSENSUS (Investing.com / Bloomberg / Argaam — May 2026) ─────────
# named[] = individual broker targets where name could be identified
ANALYST_CONSENSUS = {
    "الراجحي": {
        "avg": 113.5, "high": 126.0, "low": 99.4, "count": 8,
        "rating": "شراء", "source_date": "مايو 2026",
        "named": [
            {"house": "الجزيرة كابيتال",      "target": 102.0, "rec": "زيادة المراكز (OW)"},
            {"house": "مورغان ستانلي",         "target": 94.5,  "rec": "محايد (Neutral)"},
            {"house": "إجماع (8 محللين)",      "target": 113.5, "rec": "شراء (Buy)"},
            {"house": "الأعلى (High)",          "target": 126.0, "rec": "شراء قوي"},
            {"house": "الأدنى (Low)",           "target": 99.4,  "rec": "شراء"},
        ]
    },
    "الإنماء": {
        "avg": 27.6, "high": 41.0, "low": 25.0, "count": 8,
        "rating": "شراء", "source_date": "مايو 2026",
        "named": [
            {"house": "إجماع (8 محللين)",      "target": 27.6,  "rec": "شراء (Buy)"},
            {"house": "الأعلى (High)",          "target": 41.0,  "rec": "شراء قوي"},
            {"house": "الأدنى (Low)",           "target": 25.0,  "rec": "محايد"},
        ]
    },
    "STC": {
        "avg": 47.7, "high": 55.0, "low": 41.1, "count": 8,
        "rating": "شراء", "source_date": "مايو 2026",
        "named": [
            {"house": "إجماع (8 محللين)",      "target": 47.7,  "rec": "شراء (Buy)"},
            {"house": "الأعلى (High)",          "target": 55.0,  "rec": "شراء قوي"},
            {"house": "الأدنى (Low)",           "target": 41.1,  "rec": "محايد"},
        ]
    },
    "سال": {
        "avg": 178.4, "high": 196.1, "low": 165.0, "count": None,
        "rating": "محايد", "source_date": "مايو 2026",
        "named": [
            {"house": "إجماع المحللين",         "target": 178.4, "rec": "محايد (Neutral)"},
            {"house": "الأعلى (High)",          "target": 196.1, "rec": "شراء"},
            {"house": "الأدنى (Low)",           "target": 165.0, "rec": "محايد"},
        ]
    },
    "المواساة": {
        "avg": 91.6, "high": 125.0, "low": 75.5, "count": 5,
        "rating": "شراء", "source_date": "مايو 2026",
        "named": [
            {"house": "إجماع (5+ محللين)",     "target": 91.6,  "rec": "شراء (Buy)"},
            {"house": "الأعلى (High)",          "target": 125.0, "rec": "شراء قوي"},
            {"house": "الأدنى (Low)",           "target": 75.5,  "rec": "محايد"},
        ]
    },
    "بدجت": {
        "avg": 90.2, "high": 100.0, "low": 80.0, "count": 8,
        "rating": "⚠ شراء قوي (قبل Q1 2026)", "source_date": "⚠ سابق لنتائج Q1 2026",
        "named": [
            {"house": "⚠ إجماع (8) — قبل انهيار Q1", "target": 90.2, "rec": "شراء قوي — يستوجب مراجعة"},
            {"house": "الأعلى (High) — قبل Q1",       "target": 100.0,"rec": "شراء قوي"},
            {"house": "الأدنى (Low) — قبل Q1",        "target": 80.0, "rec": "شراء"},
        ]
    },
    "اكسترا": {
        "avg": 111.7, "high": 140.0, "low": 93.0, "count": 8,
        "rating": "شراء قوي", "source_date": "مايو 2026",
        "named": [
            {"house": "إجماع (8 محللين)",      "target": 111.7, "rec": "شراء قوي (Strong Buy)"},
            {"house": "الأعلى (High)",          "target": 140.0, "rec": "شراء قوي"},
            {"house": "الأدنى (Low)",           "target": 93.0,  "rec": "شراء"},
        ]
    },
    "المتقدمة": {
        "avg": 38.8, "high": 50.0, "low": 31.1, "count": 8,
        "rating": "محايد", "source_date": "مايو 2026",
        "named": [
            {"house": "إجماع (8 محللين)",      "target": 38.8,  "rec": "محايد (Hold)"},
            {"house": "الأعلى (High)",          "target": 50.0,  "rec": "شراء — عند التعافي"},
            {"house": "الأدنى (Low)",           "target": 31.1,  "rec": "بيع / تخفيض"},
        ]
    },
    "بنيان ريت": {
        "avg": None, "high": None, "low": None, "count": None,
        "rating": "غير متاح", "source_date": "غير متاح",
        "named": [
            {"house": "لا تغطية موثقة من بيوت خبرة", "target": None, "rec": "—"},
        ]
    },
}

# ─── EXTRA METRICS FOR ADVANCED VALUATION ─────────────────────────────────────
# DCF assumptions per company: wacc, fcf_reinvestment_rate, st_growth, lt_growth
DCF_PARAMS = {
    "الراجحي":   {"wacc": 0.105, "reinv": 0.15, "g_st": 0.08, "g_lt": 0.03, "years": 5, "use_dps": True},
    "الإنماء":   {"wacc": 0.105, "reinv": 0.15, "g_st": 0.10, "g_lt": 0.03, "years": 5, "use_dps": True},
    "STC":       {"wacc": 0.095, "reinv": 0.35, "g_st": 0.06, "g_lt": 0.03, "years": 5, "use_dps": False},
    "سال":       {"wacc": 0.110, "reinv": 0.25, "g_st": 0.12, "g_lt": 0.03, "years": 5, "use_dps": False},
    "المواساة":  {"wacc": 0.100, "reinv": 0.20, "g_st": 0.06, "g_lt": 0.03, "years": 5, "use_dps": False},
    "بدجت":      {"wacc": 0.115, "reinv": 0.40, "g_st": 0.07, "g_lt": 0.03, "years": 5, "use_dps": False},
    "اكسترا":    {"wacc": 0.115, "reinv": 0.15, "g_st": 0.10, "g_lt": 0.03, "years": 5, "use_dps": False},
    "المتقدمة":  {"wacc": 0.130, "reinv": 0.30, "g_st": 0.05, "g_lt": 0.03, "years": 5, "use_dps": False},
    "بنيان ريت": {"wacc": 0.085, "reinv": 0.05, "g_st": 0.04, "g_lt": 0.03, "years": 5, "use_dps": False},
}

# EV/EBITDA sector benchmarks (x)
EV_EBITDA_BENCH = {
    "الراجحي": None, "الإنماء": None,   # Banks → N/A
    "STC": 8.0, "سال": 12.0, "المواساة": 14.0,
    "بدجت": 8.0, "اكسترا": 10.0, "المتقدمة": 8.0,
    "بنيان ريت": None,  # REIT → use FFO
}

def compute_dcf(company_data, params):
    """Simple 5-year DCF. For banks uses DPS stream (DDM variant)."""
    shares = company_data["shares_m"]
    wacc = params["wacc"]
    g_st = params["g_st"]
    g_lt = params["g_lt"]
    reinv = params["reinv"]
    years = params["years"]

    if company_data["profits_m"].get(2024, 0) <= 0:
        return None  # Losses — DCF unreliable

    if params["use_dps"]:
        # DDM variant for banks: use latest DPS as base
        annual = get_annual_divs(company_data)
        base = annual.get(2024) or annual.get(2023)
        if not base:
            return None
    else:
        ocf = company_data["op_cashflow_m"].get(2024, 0)
        base_total = ocf * (1 - reinv)
        base = base_total / shares  # per share

    pv_sum = 0.0
    fcf = base
    for t in range(1, years + 1):
        fcf *= (1 + g_st)
        pv_sum += fcf / (1 + wacc) ** t

    # Terminal value (Gordon)
    fcf_terminal = base * (1 + g_st) ** years * (1 + g_lt)
    terminal_pv = (fcf_terminal / (wacc - g_lt)) / (1 + wacc) ** years
    return round(pv_sum + terminal_pv, 1)

def compute_ev_ebitda_fv(company_data, bench):
    """EV/EBITDA fair value: EBITDA * sector_bench / shares."""
    if bench is None:
        return None
    ocf = company_data["op_cashflow_m"].get(2024, 0)
    ebitda = ocf * 1.12  # rough EBITDA ≈ OCF / (1 - D&A effect)
    shares = company_data["shares_m"]
    if ebitda <= 0 or shares <= 0:
        return None
    return round(ebitda * bench / shares, 1)

def compute_peg(company_data):
    """PEG = P/E ÷ 3-yr EPS CAGR %."""
    pe = company_data.get("pe")
    if not pe:
        return None
    profits = company_data["profits_m"]
    p_start = profits.get(2021)
    p_end = profits.get(2024)
    if not p_start or not p_end or p_start <= 0 or p_end <= 0:
        return None
    cagr = (p_end / p_start) ** (1/3) - 1
    if cagr <= 0:
        return None
    return round(pe / (cagr * 100), 2)

def compute_ps(company_data):
    """Price/Sales ratio."""
    price = company_data["price_sar"]
    shares = company_data["shares_m"]
    rev = company_data["revenues_m"].get(2024)
    if not rev or rev <= 0:
        return None
    market_cap = price * shares
    return round(market_cap / rev, 1)

def get_annual_divs(company_data):
    """Compute annual dividend totals from the divs list."""
    annual = {}
    for (period, pct_str, dps, total_m, note, _) in company_data["divs"]:
        year = None
        for y in [2020, 2021, 2022, 2023, 2024, 2025, 2026]:
            if str(y) in period:
                year = y
                break
        if year and dps is not None:
            annual[year] = annual.get(year, 0) + dps
    return annual


# ─── METHODOLOGY SHEET ────────────────────────────────────────────────────────
def write_methodology_sheet(wb):
    ws = wb.create_sheet("📚 منهجية التقييم")
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "دليل أدوات التقييم — ما تقيسه كل أداة، الصيغة، وأنسب الشركات لها"
    c.font = Font(name='Calibri', size=13, bold=True, color=C_WHITE)
    c.fill = fill(C_DARK_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 32

    cols = ["الأداة", "الصيغة", "ما تقيسه", "الأنسب لنوع الشركات", "المحدودية", "Benchmark السعودي", "الشركات المُستخدمة معها"]
    widths = [22, 30, 38, 34, 36, 28, 36]
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = hdr_font(10)
        c.fill = fill(C_SUB_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True, readingOrder=2)
        c.border = make_border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 30

    METHODS = [
        {
            "name": "P/E — نسبة السعر إلى الأرباح",
            "formula": "سعر السهم ÷ ربح السهم (EPS)",
            "measures": "كم يدفع المستثمر مقابل كل ريال من الأرباح — مقياس الغلاء/الرخص النسبي",
            "best_for": "✅ شركات ناضجة ذات أرباح مستقرة وقابلة للتنبؤ",
            "limits": "❌ لا يعمل مع الخسائر | يتأثر بالبنود غير المتكررة | يتجاهل هيكل رأس المال",
            "benchmark": "TASI: متوسط 15-18x | بنوك: 12-16x | تجزئة: 14-20x",
            "used_for": "الراجحي، الإنماء، المواساة، بدجت، اكسترا",
            "color": "EBF5FB",
        },
        {
            "name": "P/B — السعر إلى القيمة الدفترية",
            "formula": "سعر السهم ÷ (حقوق الملكية ÷ عدد الأسهم)",
            "measures": "هل يدفع السوق علاوة على صافي أصول الشركة؟ — مقارنة بالقيمة المحاسبية",
            "best_for": "✅ البنوك والشركات المالية والعقارية حيث الأصول الملموسة هي المحرك",
            "limits": "❌ يتجاهل الأصول غير الملموسة (العلامة التجارية) | يُضلل في الصناعات كثيفة المعرفة",
            "benchmark": "بنوك سعودية: 2-4x | REITs: 0.9-1.2x | صناعة: 1.5-3x",
            "used_for": "الراجحي، الإنماء، بنيان ريت، المتقدمة",
            "color": "E8F8F5",
        },
        {
            "name": "DDM — نموذج خصم التوزيعات (Gordon Growth)",
            "formula": "DPS₁ ÷ (WACC − g)  حيث g = معدل نمو التوزيعات",
            "measures": "القيمة الحالية لتيار التوزيعات المستقبلي — يعكس القيمة للمستثمر الباحث عن الدخل",
            "best_for": "✅ شركات ذات سياسة توزيع ثابتة ومستدامة وأرباح مستقرة",
            "limits": "❌ لا ينطبق على الشركات غير الموزعة | حساس جداً لافتراض g | يهمل الاحتفاظ بالأرباح",
            "benchmark": "WACC: 9-11% | g: 3-5% | تعطي قيمة أعلى مع خفض الفائدة",
            "used_for": "الراجحي، الإنماء، المواساة، اكسترا، بنيان ريت",
            "color": "FEF9E7",
        },
        {
            "name": "DCF — خصم التدفقات النقدية الحرة",
            "formula": "Σ [FCF_t/(1+WACC)^t] + Terminal Value/(1+WACC)^n\nFCF = التدفق التشغيلي × (1 − معدل الاستثمار)",
            "measures": "القيمة الجوهرية المطلقة بناءً على التدفقات النقدية الفعلية المتوقعة — الأشمل نظرياً",
            "best_for": "✅ جميع الشركات المولدة للتدفقات | الأمثل للشركات ذات النمو العالي",
            "limits": "❌ حساس جداً للافتراضات (WACC، g، CAPEX) | لا يصلح للخسائر | أرقامه أقل موثوقية للمبتدئين",
            "benchmark": "WACC: بنوك 10.5% | اتصالات 9.5% | تجزئة 11.5% | بتروكيماويات 13% | g طويل: 3%",
            "used_for": "STC، سال، المواساة، بدجت، اكسترا، بنيان ريت (FFO-based)",
            "color": "F9EBEA",
        },
        {
            "name": "EV/EBITDA — قيمة المنشأة إلى EBITDA",
            "formula": "EV ÷ EBITDA  حيث EV = القيمة السوقية + الديون − النقد\nEBITDA = أرباح قبل الفوائد والضرائب والاستهلاك",
            "measures": "التقييم مستقلاً عن هيكل رأس المال والسياسة الضريبية والاستهلاك — مقارنة دولية",
            "best_for": "✅ شركات التشغيل بديون عالية (تأجير، اتصالات) | المقارنات القطاعية الدولية",
            "limits": "❌ لا يصلح للبنوك | يتطلب بيانات تفصيلية عن الديون | يهمل فرص النمو",
            "benchmark": "اتصالات: 7-10x | لوجستيات: 10-14x | رعاية صحية: 12-16x | تجزئة: 8-12x",
            "used_for": "STC، سال، المواساة، بدجت، اكسترا، المتقدمة",
            "color": "F5EEF8",
        },
        {
            "name": "PEG — نسبة النمو إلى P/E",
            "formula": "P/E ÷ معدل نمو EPS السنوي (%)\nمثال: P/E=15 ونمو=20% → PEG=0.75",
            "measures": "هل P/E المرتفع مبرر بمعدل النمو؟ — يُعادل قيمة الشركات بمختلف مستويات النمو",
            "best_for": "✅ شركات النمو العالي | مقارنة شركتين بـP/E مختلف | تقييم الجزء النمو في المحفظة",
            "limits": "❌ يعتمد على دقة تقديرات النمو | لا يصلح مع الخسائر | يُضلل مع النمو المؤقت",
            "benchmark": "PEG < 1x = جذاب جداً | 1-1.5x = معقول | > 2x = مكلف | < 0.5x = فرصة نادرة",
            "used_for": "الإنماء، سال، المواساة، بدجت، اكسترا",
            "color": "EAFAF1",
        },
        {
            "name": "P/S — السعر إلى الإيرادات",
            "formula": "القيمة السوقية الكلية ÷ إجمالي الإيرادات السنوية",
            "measures": "مضاعف المبيعات — مفيد عند غياب الأرباح أو تشويهها ببنود استثنائية",
            "best_for": "✅ شركات النمو في مرحلة مبكرة | شركات ذات أرباح متقلبة | مقارنة قطاع التجزئة",
            "limits": "❌ يتجاهل كلياً الربحية والهامش | شركة بمبيعات عالية وخسائر تبدو رخيصة زيفاً",
            "benchmark": "تجزئة إلكترونيات: 0.5-2x | رعاية صحية: 2-5x | اتصالات: 1-3x | لوجستيات: 3-8x",
            "used_for": "اكسترا، سال، المتقدمة (خسائر P/E)",
            "color": "FEF5E7",
        },
        {
            "name": "P/FFO — السعر إلى تدفق العمليات (REITs فقط)",
            "formula": "سعر الوحدة ÷ FFO للوحدة\nFFO = صافي الدخل + الاستهلاك − مكاسب بيع الأصول",
            "measures": "الأداة الصحيحة لتقييم REITs — يستبعد الاستهلاك الذي يُضخم الخسارة المحاسبية للعقارات",
            "best_for": "✅ صناديق الاستثمار العقاري (REITs) حصراً | يعكس القدرة الفعلية على التوزيع",
            "limits": "❌ لا ينطبق على غير REITs | يتفاوت تعريف FFO بين الشركات",
            "benchmark": "REITs السعودية: 12-18x | REITs الأمريكية: 15-22x | عائد توزيع ≥ 5%",
            "used_for": "بنيان ريت (4340) حصراً",
            "color": "E8EAF6",
        },
    ]

    for r, m in enumerate(METHODS, 3):
        bg = m["color"]
        row_data = [m["name"], m["formula"], m["measures"], m["best_for"],
                    m["limits"], m["benchmark"], m["used_for"]]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill(bg)
            c.border = make_border()
            c.alignment = Alignment(horizontal='right', vertical='center',
                                    wrap_text=True, readingOrder=2)
            bold = col == 1
            c.font = Font(name='Calibri', size=10, bold=bold, color="1A1A2E")
        ws.row_dimensions[r].height = 58

    # Summary row
    sr = len(METHODS) + 3
    ws.merge_cells(f"A{sr}:G{sr}")
    c = ws[f"A{sr}"]
    c.value = ("⚖ مبدأ التقييم: لا توجد أداة مثالية واحدة. الجمع بين 3+ طرق يُعطي نطاقاً للقيمة العادلة أكثر موثوقية."
               " | عندما تتقارب 3 طرق مختلفة على نفس المستوى السعري — تكون الثقة عالية في ذلك التقدير.")
    c.font = Font(name='Calibri', size=10, bold=True, color=C_WHITE)
    c.fill = fill(C_DARK_NAVY)
    c.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws.row_dimensions[sr].height = 36


# ─── ANALYST HOUSE FORECASTS SHEET ───────────────────────────────────────────
def write_analyst_sheet(wb, companies, analyst_data):
    ws = wb.create_sheet("🏦 تقييمات المحللين")
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = ("تقييمات بيوت الخبرة والإجماع لـ 9 شركات  |  المصدر: Investing.com · Bloomberg · Argaam · الجزيرة كابيتال  |  مايو 2026")
    c.font = Font(name='Calibri', size=12, bold=True, color=C_WHITE)
    c.fill = fill(C_DARK_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = ("⚠ ملاحظة: أسعار بيوت الخبرة المسماة محدودة الإتاحة للعموم — البيانات المُتحقق منها تعتمد على إجماع Investing.com/Bloomberg."
               " | ⚠ أهداف بدجت مُحتسبة قبل نتائج Q1 2026 المخيبة للآمال.")
    c.font = Font(name='Calibri', size=9, italic=True, color="7D6608")
    c.fill = fill("FEF9E7")
    c.alignment = Alignment(horizontal='right', wrap_text=True, vertical='center')
    ws.row_dimensions[2].height = 32

    hdrs = ["الشركة", "الرمز", "السعر الحالي\n(ريال)", "بيت الخبرة / المصدر",
            "السعر المستهدف\n(ريال)", "التوصية", "Upside%"]
    widths = [18, 9, 16, 38, 18, 36, 14]
    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = hdr_font(10)
        c.fill = fill(C_SUB_HEADER)
        c.alignment = center()
        c.border = make_border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 32

    row = 4
    SECTOR_COLORS = {
        "الراجحي": "D6EAF8", "الإنماء": "D6EAF8", "STC": "E8DAEF",
        "سال": "D5F5E3", "المواساة": "FADBD8", "بدجت": "FDEBD0",
        "اكسترا": "D5F5E3", "المتقدمة": "FDFEFE", "بنيان ريت": "EAF2FF",
    }

    for name in COMPANY_ORDER:
        d = companies[name]
        ac = analyst_data.get(name, {})
        named_list = ac.get("named", [])
        price = d["price_sar"]
        ticker = d["ticker"]
        bg = SECTOR_COLORS.get(name, "F2F3F4")

        # Consensus summary row (merged company name)
        start_row = row
        for entry in named_list:
            target = entry.get("target")
            rec = entry.get("rec", "—")
            house = entry.get("house", "—")

            upside_str = "—"
            upside_color = "000000"
            if target:
                upside = (target - price) / price * 100
                upside_str = f"{upside:+.1f}%"
                upside_color = C_POSITIVE if upside > 5 else (C_NEGATIVE if upside < -5 else C_WARNING)

            # Company & ticker only on first row
            if row == start_row:
                c = ws.cell(row=row, column=1, value=name)
                c.font = Font(name='Calibri', size=10, bold=True, color="1A1A2E")
                c.fill = fill(bg); c.alignment = center(); c.border = make_border()
                c = ws.cell(row=row, column=2, value=ticker)
                c.font = Font(name='Calibri', size=9, color="555555")
                c.fill = fill(bg); c.alignment = center(); c.border = make_border()
                c = ws.cell(row=row, column=3, value=price)
                c.font = Font(name='Calibri', size=10, bold=True, color=C_DARK_NAVY)
                c.fill = fill(bg); c.alignment = center(); c.number_format = '#,##0.00'
                c.border = make_border()
            else:
                for col in [1, 2, 3]:
                    c = ws.cell(row=row, column=col, value="")
                    c.fill = fill(bg); c.border = make_border()

            c = ws.cell(row=row, column=4, value=house)
            c.font = Font(name='Calibri', size=10, color="1A1A2E")
            c.fill = fill(bg); c.alignment = right_align(); c.border = make_border()

            c = ws.cell(row=row, column=5, value=target if target else "—")
            c.font = Font(name='Calibri', size=11, bold=True,
                         color=C_POSITIVE if target and target > price else (C_NEGATIVE if target and target < price else "555555"))
            c.fill = fill(bg); c.alignment = center()
            c.number_format = '#,##0.00'; c.border = make_border()

            c = ws.cell(row=row, column=6, value=rec)
            c.font = Font(name='Calibri', size=9, color="1A1A2E")
            c.fill = fill(bg); c.alignment = center(); c.border = make_border()

            c = ws.cell(row=row, column=7, value=upside_str)
            c.font = Font(name='Calibri', size=10, bold=True, color=upside_color)
            c.fill = fill(bg); c.alignment = center(); c.border = make_border()

            ws.row_dimensions[row].height = 20
            row += 1

        # Consensus summary bar
        avg = ac.get("avg")
        low = ac.get("low")
        high = ac.get("high")
        count = ac.get("count")
        rating = ac.get("rating", "—")
        src = ac.get("source_date", "—")
        summary = (f"الإجماع: متوسط {avg} | نطاق {low}–{high} | {count or '?'} محللين | {rating} | {src}" if avg
                   else "لا يوجد إجماع موثق")
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]
        c.value = summary
        c.font = Font(name='Calibri', size=9, italic=True, color=C_WHITE)
        c.fill = fill(C_SUB_HEADER)
        c.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[row].height = 16
        row += 1

    # Upside legend
    ws.merge_cells(f"A{row+1}:G{row+1}")
    c = ws[f"A{row+1}"]
    c.value = "🟢 Upside > +5%  =  فرصة رأسمالية   |   🟡 بين ±5%  =  تقييم عادل   |   🔴 Downside > -5%  =  مكلف أو بلغ هدفه"
    c.font = Font(name='Calibri', size=9, bold=True, color=C_DARK_NAVY)
    c.fill = fill("F8F9FA"); c.alignment = Alignment(horizontal='right')
    ws.row_dimensions[row+1].height = 18


def write_overview_sheet(wb, companies):
    ws = wb.create_sheet("📊 نظرة عامة")
    ws.sheet_view.rightToLeft = True

    # Title
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "دراسة استثمارية - المحفظة السعودية  |  مصدر البيانات: تداول · أرقام · مباشر  |  مايو 2026"
    c.font = Font(name='Calibri', size=14, bold=True, color=C_WHITE)
    c.fill = fill(C_DARK_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 30

    # Legend row
    ws.merge_cells("A2:N2")
    c = ws["A2"]
    c.value = "التصنيف: 🟢 Value = قيمة   🔵 Growth = نمو   🟡 Blend = مزيج      ⚠ = ملاحظة تحقق"
    c.font = Font(name='Calibri', size=9, italic=True, color="555555")
    c.alignment = center()

    # Headers
    headers = [
        "الشركة", "الرمز", "القطاع", "التصنيف",
        "السعر\n(ريال)", "م.ربحية\n(P/E)", "م.دفتري\n(P/B)",
        "عائد\nتوزيع%", "ROE%",
        "ربح 2024\n(م.ريال)", "EPS\n2024",
        "توزيع/سهم\n2024",
        "نسبة توزيع\n2024%",
        "التوصية",
    ]
    row = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hdr_font(10)
        c.fill = fill(C_DARK_NAVY)
        c.alignment = center()
        c.border = make_border()
    ws.row_dimensions[row].height = 38

    # Data rows
    CLASS_COLOR = {"Value": C_VALUE[:-2] if len(C_VALUE) > 6 else C_VALUE,
                   "Growth": "FFF3CD", "Blend": "D6EAF8"}
    RECOM_COLOR = {"تراكم بقوة": C_POSITIVE, "تراكم": "1A6B30",
                   "احتفاظ / تراكم عند الضعف": C_NEUTRAL,
                   "احتفاظ للدخل": C_NEUTRAL,
                   "مراقبة / تراكم تدريجي": C_WARNING,
                   "مراقبة - انتظار توضيح Q1 2026": C_NEGATIVE,
                   "مضاربة / وزن محدود": C_NEGATIVE}

    for r, name in enumerate(COMPANY_ORDER, 4):
        d = companies[name]
        shares = d["shares_m"]
        profit_2024 = d["profits_m"].get(2024, 0)
        eps_2024 = profit_2024 / shares if shares and profit_2024 else None

        annual_divs = get_annual_divs(d)
        dps_2024 = annual_divs.get(2024, 0)
        payout_2024 = (dps_2024 / eps_2024 * 100) if (eps_2024 and eps_2024 > 0 and dps_2024) else None

        cl_bg = {"Value": "F0FFF0", "Growth": "FFF8E1", "Blend": "EAF2FF"}.get(d["classification"], C_WHITE)
        row_bg = cl_bg

        def wc(col, val, fmt=None, bold=False, color="000000"):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name='Calibri', size=10, bold=bold, color=color)
            cell.fill = fill(row_bg)
            cell.alignment = center()
            cell.border = make_border()
            if fmt:
                cell.number_format = fmt
            return cell

        wc(1, name, bold=True)
        wc(2, d["ticker"])
        wc(3, d["sector"])

        cl_map = {"Value": "🟢 Value", "Growth": "🔵 Growth", "Blend": "🟡 Blend"}
        wc(4, cl_map.get(d["classification"], d["classification"]))

        wc(5, d["price_sar"], '#,##0.00')
        pe_val = d.get("pe")
        wc(6, pe_val if pe_val else "خسارة", '#,##0.0' if pe_val else None)
        wc(7, d.get("pb"), '#,##0.00')
        wc(8, d.get("div_yield_pct"), '0.0%' if False else '#,##0.0')
        roe = d.get("roe_pct")
        wc(9, roe if roe else "-", '#,##0.0' if roe else None)
        wc(10, profit_2024, '#,##0')
        wc(11, round(eps_2024, 2) if eps_2024 else "-", '#,##0.00' if eps_2024 else None)
        wc(12, dps_2024 if dps_2024 else "-", '#,##0.00' if dps_2024 else None)
        payout_str = f"{payout_2024:.1f}%" if payout_2024 else ("-" if profit_2024 < 0 else "N/A")
        wc(13, payout_str)

        rec = d.get("recommendation", "")
        c_rec = ws.cell(row=r, column=14, value=rec)
        rec_color = RECOM_COLOR.get(rec, C_NEUTRAL)
        c_rec.font = Font(name='Calibri', size=10, bold=True, color=C_WHITE)
        c_rec.fill = fill(rec_color)
        c_rec.alignment = center()
        c_rec.border = make_border()
        ws.row_dimensions[r].height = 22

    # Source note
    note_row = 4 + len(COMPANY_ORDER) + 1
    ws.merge_cells(f"A{note_row}:N{note_row}")
    c = ws[f"A{note_row}"]
    c.value = ("⚠ تنبيهات: (1) الراجحي: منح سهم لكل سهمين في H2 2025 يُعدل EPS/DPS لاحقاً. "
               "(2) المتقدمة: خسارة 2024، مستثناة من P/E. "
               "(3) بدجت: انخفاض حاد في Q1 2026 يستوجب مراقبة. "
               "(4) STC: ربح 2024 يشمل بنود غير متكررة. "
               "(5) جميع الأرقام بالريال السعودي.")
    c.font = Font(name='Calibri', size=9, italic=True, color="8B0000")
    c.fill = fill("FFF9C4")
    c.alignment = Alignment(horizontal='right', wrap_text=True, vertical='center')
    ws.row_dimensions[note_row].height = 40

    # Column widths
    col_widths = [16, 8, 14, 14, 10, 10, 10, 10, 8, 14, 10, 14, 12, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_dividend_sheet(wb, companies):
    ws = wb.create_sheet("💰 التوزيعات النقدية")
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "تحليل التوزيعات النقدية التاريخية - مقارنة الأرباح الموزعة بصافي الربح والتدفق النقدي التشغيلي"
    c.font = hdr_font(12)
    c.fill = fill(C_DARK_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    current_row = 2
    for name in COMPANY_ORDER:
        d = companies[name]
        # Company header
        ws.merge_cells(f"A{current_row}:G{current_row}")
        c = ws[f"A{current_row}"]
        c.value = f"  {name} ({d['ticker']}) — {d['sector']}  {d.get('note_capital', '')}"
        c.font = Font(name='Calibri', size=11, bold=True, color=C_WHITE)
        c.fill = fill(C_SUB_HEADER)
        c.alignment = left_align()
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Sub headers
        sub_hdrs = ["الفترة", "النسبة %", "توزيع/سهم (ريال)", "إجمالي التوزيع (م.ريال)",
                    "صافي الربح المقارن (م.ريال)", "نسبة التوزيع %", "ملاحظة"]
        for col, h in enumerate(sub_hdrs, 1):
            c = ws.cell(row=current_row, column=col, value=h)
            c.font = hdr_font(9)
            c.fill = fill("2E4172")
            c.alignment = center()
            c.border = make_border()
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Dividend rows
        profits_by_year = d["profits_m"]
        for (period, pct_str, dps, total_m, note, _) in d["divs"]:
            year = None
            for y in [2020, 2021, 2022, 2023, 2024, 2025, 2026]:
                if str(y) in period:
                    year = y
                    break
            net_profit = profits_by_year.get(year, None) if year else None

            if dps is not None and net_profit and net_profit > 0 and total_m:
                payout = total_m / net_profit * 100
                payout_str = f"{payout:.1f}%"
                payout_color = C_POSITIVE if payout <= 80 else (C_WARNING if payout <= 100 else C_NEGATIVE)
            else:
                payout_str = "—"
                payout_color = C_NEUTRAL

            row_data = [period, pct_str, dps if dps is not None else "—",
                        total_m if total_m else "—",
                        net_profit if net_profit else "—",
                        payout_str,
                        note if note else ""]

            bg = C_LIGHT_GREEN if (dps and dps > 0) else C_LIGHT_RED if dps == 0 else C_LIGHT_GRAY
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=current_row, column=col, value=val)
                c.font = body_font(9)
                c.fill = fill(bg)
                c.border = make_border("DDDDDD")
                c.alignment = center() if col > 1 else left_align()
                if col == 6 and payout_str not in ["—", "N/A"]:
                    c.font = Font(name='Calibri', size=9, bold=True, color=payout_color)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        # Annual summary row
        annual_divs = get_annual_divs(d)
        ws.merge_cells(f"A{current_row}:G{current_row}")
        c = ws[f"A{current_row}"]
        years_summary = " | ".join([f"{y}: {v:.2f} ريال/سهم" for y, v in sorted(annual_divs.items()) if v > 0])
        c.value = f"  ملخص سنوي: {years_summary}"
        c.font = Font(name='Calibri', size=9, italic=True, color="1A1A1A")
        c.fill = fill("FFFDE7")
        c.alignment = left_align()
        ws.row_dimensions[current_row].height = 18
        current_row += 2

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 32


def write_financials_sheet(wb, companies):
    ws = wb.create_sheet("📈 القوائم المالية")
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "القوائم المالية 2021-2024 (جميع الأرقام بالمليون ريال سعودي)"
    c.font = hdr_font(12)
    c.fill = fill(C_DARK_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    current_row = 2
    years = [2021, 2022, 2023, 2024]

    for name in COMPANY_ORDER:
        d = companies[name]
        ws.merge_cells(f"A{current_row}:G{current_row}")
        c = ws[f"A{current_row}"]
        c.value = f"  {name} ({d['ticker']}) — {d['sector']}"
        c.font = Font(name='Calibri', size=11, bold=True, color=C_WHITE)
        c.fill = fill(C_SUB_HEADER)
        c.alignment = left_align()
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Year headers
        ws.cell(row=current_row, column=1, value="البند").font = hdr_font(9)
        ws.cell(row=current_row, column=1).fill = fill(C_DARK_NAVY)
        ws.cell(row=current_row, column=1).alignment = center()
        for col, y in enumerate(years, 2):
            c = ws.cell(row=current_row, column=col, value=str(y))
            c.font = hdr_font(9)
            c.fill = fill(C_DARK_NAVY)
            c.alignment = center()

        # CAGR column
        cagr_col = len(years) + 2
        c = ws.cell(row=current_row, column=cagr_col, value="CAGR 3Y")
        c.font = hdr_font(9)
        c.fill = fill("2E6B8A")
        c.alignment = center()

        growth_col = cagr_col + 1
        c = ws.cell(row=current_row, column=growth_col, value="نمو 2024/2023")
        c.font = hdr_font(9)
        c.fill = fill("2E6B8A")
        c.alignment = center()

        ws.row_dimensions[current_row].height = 22
        current_row += 1

        metrics = [
            ("الإيرادات", d["revenues_m"]),
            ("صافي الربح", d["profits_m"]),
            ("التدفق النقدي التشغيلي", d["op_cashflow_m"]),
            ("حقوق المساهمين", d["equity_m"]),
        ]

        for metric_name, metric_data in metrics:
            bg = C_LIGHT_BLUE if metric_name == "الإيرادات" else (
                "FFF9C4" if metric_name == "صافي الربح" else C_LIGHT_GRAY)
            ws.cell(row=current_row, column=1, value=metric_name).font = body_font(9, bold=True)
            ws.cell(row=current_row, column=1).fill = fill(bg)
            ws.cell(row=current_row, column=1).alignment = left_align()
            ws.cell(row=current_row, column=1).border = make_border("DDDDDD")

            vals = [metric_data.get(y, None) for y in years]
            for col, v in enumerate(vals, 2):
                c = ws.cell(row=current_row, column=col, value=v)
                c.font = body_font(9, color=(C_NEGATIVE if (v is not None and v < 0) else "000000"))
                c.fill = fill(bg)
                c.alignment = center()
                c.border = make_border("DDDDDD")
                c.number_format = '#,##0'

            # CAGR
            v0, v3 = metric_data.get(2021), metric_data.get(2024)
            if v0 and v3 and v0 > 0 and v3 > 0:
                cagr = ((v3 / v0) ** (1/3) - 1) * 100
                c_cagr = ws.cell(row=current_row, column=cagr_col, value=f"{cagr:.1f}%")
                c_cagr.font = Font(name='Calibri', size=9, bold=True,
                                   color=(C_POSITIVE if cagr >= 5 else (C_WARNING if cagr >= 0 else C_NEGATIVE)))
            else:
                ws.cell(row=current_row, column=cagr_col, value="N/A")
            ws.cell(row=current_row, column=cagr_col).fill = fill(bg)
            ws.cell(row=current_row, column=cagr_col).alignment = center()
            ws.cell(row=current_row, column=cagr_col).border = make_border("DDDDDD")

            # YoY growth
            v23, v24 = metric_data.get(2023), metric_data.get(2024)
            if v23 and v24 and v23 != 0:
                yoy = (v24 - v23) / abs(v23) * 100
                yoy_str = f"{yoy:+.1f}%"
                c_yoy = ws.cell(row=current_row, column=growth_col, value=yoy_str)
                c_yoy.font = Font(name='Calibri', size=9, bold=True,
                                  color=(C_POSITIVE if yoy >= 0 else C_NEGATIVE))
            else:
                ws.cell(row=current_row, column=growth_col, value="N/A")
            ws.cell(row=current_row, column=growth_col).fill = fill(bg)
            ws.cell(row=current_row, column=growth_col).alignment = center()
            ws.cell(row=current_row, column=growth_col).border = make_border("DDDDDD")

            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 1

    ws.column_dimensions["A"].width = 26
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 14


def write_valuation_sheet(wb, companies):
    ws = wb.create_sheet("🎯 التقييم والقيمة العادلة")
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "مصفوفة التقييم الشاملة — 7 أدوات تقييم + إجماع المحللين + هامش الأمان  (انظر ورقة منهجية التقييم للشرح)"
    c.font = hdr_font(12)
    c.fill = fill(C_DARK_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    # Method legend row
    ws.merge_cells("A2:N2")
    c = ws["A2"]
    c.value = ("🟦 P/E  🟩 P/B  🟨 DDM  🟥 DCF  🟪 EV/EBITDA  🔶 P/S  🔷 PEG  |  "
               "✅ أخضر = فوق السعر الحالي (هدف أعلى)   🔴 أحمر = دون السعر الحالي")
    c.font = Font(name='Calibri', size=9, italic=True, color="2C3E50")
    c.fill = fill("EAF4FB")
    c.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[2].height = 20

    headers = [
        "الشركة", "السعر\nالحالي",
        "P/E\n(نسبة)", "P/B\n(نسبة)", "عائد\nالتوزيع%", "PEG\n(نمو/P/E)", "P/S\n(الإيرادات)",
        "قيمة عادلة\nP/E",  "قيمة عادلة\nDDM",
        "قيمة عادلة\nP/B",  "قيمة عادلة\nDCF",
        "قيمة عادلة\nEV/EBITDA",
        "إجماع\nالمحللين",
        "هامش الأمان\n(متوسط 5 طرق)%",
    ]
    col_widths = [16, 10, 9, 9, 10, 9, 9, 13, 13, 13, 13, 14, 13, 16]
    HEADER_COLORS = [
        C_DARK_NAVY, C_DARK_NAVY,
        "154360", "154360", "154360", "1A5276", "1A5276",
        "145A32", "145A32", "145A32", "7B241C", "6C3483",
        "1B4F72", C_GOLD,
    ]

    for col, (h, hc, w) in enumerate(zip(headers, HEADER_COLORS, col_widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font(9)
        c.fill = fill(hc)
        c.alignment = center()
        c.border = make_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 36

    ROW_COLORS = ["EBF5FB", "E8F8F5", "F9EBEA", "EAFAF1", "FDFEFE",
                  "FEF9E7", "EAFAF1", "F5EEF8", "EAF2FF"]

    for r_idx, name in enumerate(COMPANY_ORDER):
        d = companies[name]
        r = r_idx + 4
        price = d["price_sar"]
        bg = ROW_COLORS[r_idx % len(ROW_COLORS)]

        fv = d.get("fair_value", {})
        pe_fv  = fv.get("pe_based")
        ddm_fv = fv.get("ddm")
        pb_fv  = fv.get("pb_based")

        dcf_fv   = compute_dcf(d, DCF_PARAMS[name])
        ev_fv    = compute_ev_ebitda_fv(d, EV_EBITDA_BENCH.get(name))
        analyst_avg = ANALYST_CONSENSUS.get(name, {}).get("avg")
        peg_val  = compute_peg(d)
        ps_val   = compute_ps(d)

        # Weighted average of 5 non-analyst methods
        valids = [x for x in [pe_fv, ddm_fv, pb_fv, dcf_fv, ev_fv] if x]
        avg_fv = round(sum(valids) / len(valids), 1) if valids else None
        margin = round((avg_fv - price) / price * 100, 1) if avg_fv else None

        def wv(col, val, fmt=None, clr="1A1A2E", bold=False):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name='Calibri', size=10, bold=bold, color=clr)
            c.fill = fill(bg)
            c.alignment = center()
            c.border = make_border()
            if fmt:
                c.number_format = fmt

        def fv_cell(col, val):
            if val is None:
                wv(col, "—", clr="888888")
            else:
                clr = C_POSITIVE if val > price else C_NEGATIVE
                wv(col, val, '#,##0.0', clr, True)

        wv(1, name, bold=True)
        wv(2, price, '#,##0.00', C_DARK_NAVY, True)

        pe_v = d.get("pe")
        wv(3, pe_v if pe_v else "خسارة", clr="555555" if not pe_v else "1A1A2E")
        wv(4, d.get("pb"), '#,##0.1')
        wv(5, d.get("div_yield_pct"), '#,##0.0')

        if peg_val:
            peg_clr = C_POSITIVE if peg_val < 1.0 else (C_WARNING if peg_val < 2.0 else C_NEGATIVE)
            wv(6, peg_val, '#,##0.00', peg_clr, True)
        else:
            wv(6, "—", clr="888888")

        if ps_val:
            wv(7, ps_val, '#,##0.1')
        else:
            wv(7, "—", clr="888888")

        fv_cell(8, pe_fv)
        fv_cell(9, ddm_fv)
        fv_cell(10, pb_fv)
        fv_cell(11, dcf_fv)
        fv_cell(12, ev_fv)

        if analyst_avg:
            clr = C_POSITIVE if analyst_avg > price else C_NEGATIVE
            wv(13, analyst_avg, '#,##0.0', clr, True)
        else:
            wv(13, "—", clr="888888")

        if margin is not None:
            margin_str = f"{margin:+.1f}%"
            m_clr = C_POSITIVE if margin > 10 else (C_WARNING if margin > 0 else C_NEGATIVE)
            wv(14, margin_str, None, m_clr, True)
        else:
            wv(14, "—", clr="888888")

        ws.row_dimensions[r].height = 22

    # Notes
    note_row = 4 + len(COMPANY_ORDER) + 1
    ws.merge_cells(f"A{note_row}:N{note_row}")
    c = ws[f"A{note_row}"]
    c.value = (
        "منهجية: P/E = متوسط P/E قطاعي × EPS 2024  |  P/B = متوسط P/B قطاعي × القيمة الدفترية  |  "
        "DDM = DPS₁/(WACC−g)  |  DCF = Σ FCF/(1+WACC)^t + Terminal (5 سنوات، g_terminal=3%)  |  "
        "EV/EBITDA = EBITDA × Benchmark القطاعي ÷ الأسهم  |  "
        "PEG = P/E ÷ CAGR%  |  هامش الأمان = (متوسط 5 طرق − السعر) ÷ السعر  |  "
        "⚠ DCF وEV/EBITDA تقديرات مبنية على OCF كـProxy — راجع ورقة منهجية التقييم للتفاصيل"
    )
    c.font = Font(name='Calibri', size=8, italic=True, color="555555")
    c.fill = fill("F8F9FA")
    c.alignment = Alignment(horizontal='right', wrap_text=True, vertical='center')
    ws.row_dimensions[note_row].height = 42


def write_portfolio_sheet(wb, companies):
    ws = wb.create_sheet("🏦 تكوين المحفظة")
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "توصية تكوين المحفظة — محفظة متوازنة بين القيمة والنمو مع 15-20% كاش"
    c.font = hdr_font(12)
    c.fill = fill(C_DARK_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    # Allocation suggestion
    allocations = [
        ("الراجحي",   "Blend",  18, "نواة البنوك الإسلامية · توزيعات مستقرة · رؤية 2030",        "منخفض-متوسط"),
        ("STC",        "Blend",  15, "أعلى عائد توزيع · البنية الرقمية الوطنية",                  "منخفض"),
        ("اكسترا",    "Value",  12, "قطاع التجزئة الدفاعي · توزيعات ممتازة · هامش أمان جيد",    "منخفض-متوسط"),
        ("المواساة",  "Value",  10, "رعاية صحية دفاعية · توزيع سنوي منتظم",                       "منخفض"),
        ("الإنماء",   "Growth", 10, "أسرع البنوك نمواً · تراجع تقييم يوفر دخولاً جيدة",          "منخفض-متوسط"),
        ("سال",        "Growth",  7, "الاستفادة من نمو قطاع الخدمات اللوجستية الجوية",             "متوسط"),
        ("بنيان ريت", "Value",   8, "دخل إيجاري ثابت · عائد توزيع مرتفع · تنويع القطاعات",       "منخفض-متوسط"),
        ("المتقدمة",  "Blend",   5, "وزن محدود · مضاربة على العودة للربحية + رؤية 2030",          "مرتفع"),
        ("بدجت",      "Growth",  0, "⚠ تأجيل حتى توضيح أسباب تراجع Q1 2026",                   "متوسط-مرتفع"),
        ("كاش",        "—",      15, "احتياطي للشراء عند الانخفاض · الحد الأدنى المقترح",         "صفر"),
    ]

    headers = ["الشركة / الأصل", "التصنيف", "الوزن المقترح %", "مبلغ مقترح (100K)", "المبرر", "مستوى المخاطرة"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hdr_font(9)
        c.fill = fill(C_DARK_NAVY)
        c.alignment = center()
        c.border = make_border()
    ws.row_dimensions[2].height = 28

    for r, (name, cls, wt, rationale, risk) in enumerate(allocations, 3):
        risk_colors = {"منخفض": C_POSITIVE, "منخفض-متوسط": "2E7D32",
                       "متوسط": C_WARNING, "متوسط-مرتفع": "BF360C",
                       "مرتفع": C_NEGATIVE, "صفر": C_NEUTRAL}
        cls_colors = {"Value": "1B5E20", "Growth": "E65100", "Blend": "1A237E", "—": "455A64"}

        bg = "ECEFF1" if name == "كاش" else ("FFF3E0" if wt == 0 else (
            "F0FFF0" if cls == "Value" else ("FFF8E1" if cls == "Growth" else "EAF2FF")))

        def wc(col, val, clr="000000", bold=False):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name='Calibri', size=10, bold=bold, color=clr)
            c.fill = fill(bg)
            c.alignment = center()
            c.border = make_border()

        wc(1, name, bold=True)
        wc(2, cls, cls_colors.get(cls, "000000"), True)
        wc(3, f"{wt}%", C_POSITIVE if wt >= 10 else (C_NEUTRAL if wt >= 5 else C_WARNING), True)
        amount = wt * 1000 if wt > 0 else 0
        wc(4, f"{amount:,} ريال" if amount > 0 else "—")
        c = ws.cell(row=r, column=5, value=rationale)
        c.font = Font(name='Calibri', size=9, color="333333")
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        c.border = make_border()
        wc(6, risk, risk_colors.get(risk, "000000"), True)
        ws.row_dimensions[r].height = 22

    # Strategy notes
    notes_row = 3 + len(allocations) + 1
    ws.merge_cells(f"A{notes_row}:H{notes_row+5}")
    c = ws[f"A{notes_row}"]
    c.value = """استراتيجية الشراء التدريجي (Staggered Entry):
• المرحلة 1 (الآن): الراجحي + STC + اكسترا + المواساة = 55% من المحفظة (على دفعتين)
• المرحلة 2 (عند الضعف أو بعد نتائج Q2 2026): الإنماء + بنيان ريت + سال = 25%
• المرحلة 3 (عند التأكد من انتعاش): المتقدمة = 5% (مضاربة)
• بدجت: انتظار نتائج Q2 2026 لفهم أسباب تراجع Q1 2026 قبل أي قرار
• الكاش 15-20%: للاستفادة من أي تراجع حاد في السوق
• إعادة توازن: كل 6 أشهر أو عند تجاوز أي ورقة للوزن بـ 5%"""
    c.font = Font(name='Calibri', size=10)
    c.fill = fill("E8F5E9")
    c.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    c.border = make_border("A5D6A7")
    for nr in range(notes_row, notes_row + 6):
        ws.row_dimensions[nr].height = 20

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 18


def write_company_sheets(wb, companies):
    for name in COMPANY_ORDER:
        d = companies[name]
        ws = wb.create_sheet(f"{d['ticker']}-{name[:4]}")
        ws.sheet_view.rightToLeft = True

        # Title
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = f"{name} ({d['ticker']}) — {d['sector']}"
        c.font = hdr_font(13)
        c.fill = fill(C_DARK_NAVY)
        c.alignment = center()
        ws.row_dimensions[1].height = 30

        # Key metrics box
        metrics_box = [
            ("السعر الحالي (ريال)", d["price_sar"]),
            ("رأس المال (م.ريال)", d["capital_m"]),
            ("عدد الأسهم (م.سهم)", d["shares_m"]),
            ("مضاعف الربحية (P/E)", d.get("pe", "خسارة")),
            ("مضاعف الدفتري (P/B)", d.get("pb")),
            ("عائد التوزيع %", d.get("div_yield_pct")),
            ("العائد على حقوق المساهمين %", d.get("roe_pct")),
            ("تصنيف المحفظة", d["classification"]),
            ("مستوى المخاطرة", d.get("risk_level")),
            ("التوصية", d.get("recommendation")),
        ]

        ws.cell(row=2, column=1, value="المؤشر").font = hdr_font(9)
        ws.cell(row=2, column=2, value="القيمة").font = hdr_font(9)
        ws.cell(row=2, column=1).fill = fill(C_SUB_HEADER)
        ws.cell(row=2, column=2).fill = fill(C_SUB_HEADER)
        ws.cell(row=2, column=1).alignment = center()
        ws.cell(row=2, column=2).alignment = center()

        for i, (k, v) in enumerate(metrics_box, 3):
            bg = C_LIGHT_GRAY if i % 2 == 0 else C_WHITE
            c = ws.cell(row=i, column=1, value=k)
            c.font = body_font(10, bold=True)
            c.fill = fill(bg)
            c.alignment = left_align()
            c.border = make_border("DDDDDD")
            c = ws.cell(row=i, column=2, value=v)
            c.font = body_font(10)
            c.fill = fill(bg)
            c.alignment = center()
            c.border = make_border("DDDDDD")

        # Profits table
        row = 14
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = "الأداء المالي التاريخي (مليون ريال)"
        c.font = hdr_font(10)
        c.fill = fill(C_SUB_HEADER)
        c.alignment = center()
        row += 1

        years = [2021, 2022, 2023, 2024]
        ws.cell(row=row, column=1, value="البند").font = hdr_font(9)
        ws.cell(row=row, column=1).fill = fill(C_DARK_NAVY)
        for ci, y in enumerate(years, 2):
            c = ws.cell(row=row, column=ci, value=str(y))
            c.font = hdr_font(9)
            c.fill = fill(C_DARK_NAVY)
            c.alignment = center()
        row += 1

        for metric_label, data_key in [("الإيرادات", "revenues_m"), ("صافي الربح", "profits_m"),
                                        ("التدفق النقدي التشغيلي", "op_cashflow_m")]:
            data = d[data_key]
            ws.cell(row=row, column=1, value=metric_label).font = body_font(9, bold=True)
            ws.cell(row=row, column=1).alignment = left_align()
            ws.cell(row=row, column=1).border = make_border("DDDDDD")
            for ci, y in enumerate(years, 2):
                v = data.get(y)
                c = ws.cell(row=row, column=ci, value=v)
                c.font = Font(name='Calibri', size=9,
                              color=(C_NEGATIVE if (v and v < 0) else "000000"))
                c.alignment = center()
                c.border = make_border("DDDDDD")
                c.number_format = '#,##0'
            row += 1

        # Vision 2030
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = "رؤية 2030 والمحفزات: " + d.get("vision2030", "")
        c.font = Font(name='Calibri', size=9, color=C_POSITIVE)
        c.fill = fill("E8F5E9")
        c.alignment = Alignment(horizontal='right', wrap_text=True)
        ws.row_dimensions[row].height = 32
        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = "المخاطر: " + d.get("key_risks", "")
        c.font = Font(name='Calibri', size=9, color=C_NEGATIVE)
        c.fill = fill("FDEDEC")
        c.alignment = Alignment(horizontal='right', wrap_text=True)
        ws.row_dimensions[row].height = 32

        # Note
        if d.get("note_capital"):
            row += 1
            ws.merge_cells(f"A{row}:F{row}")
            c = ws[f"A{row}"]
            c.value = "⚠ تنبيه: " + d["note_capital"]
            c.font = Font(name='Calibri', size=9, italic=True, color="8B0000")
            c.fill = fill("FFF9C4")
            c.alignment = Alignment(horizontal='right', wrap_text=True)
            ws.row_dimensions[row].height = 28

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14


def generate_excel(output_path):
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    write_methodology_sheet(wb)
    write_overview_sheet(wb, COMPANIES)
    write_dividend_sheet(wb, COMPANIES)
    write_financials_sheet(wb, COMPANIES)
    write_valuation_sheet(wb, COMPANIES)
    write_analyst_sheet(wb, COMPANIES, ANALYST_CONSENSUS)
    write_portfolio_sheet(wb, COMPANIES)
    write_company_sheets(wb, COMPANIES)

    wb.save(output_path)
    print(f"✅ Excel saved: {output_path}")


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(__file__), "Saudi_Portfolio_Analysis_2026.xlsx")
    generate_excel(out)
